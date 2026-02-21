import streamlit as st
import pandas as pd
import random
import re
import io
import plotly.express as px
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

# --- STREAMLIT UI SETUP ---
st.set_page_config(page_title="Smart Batcher", page_icon="📊", layout="wide")
st.title("📊 Smart Batcher - Easy Batching")

# --- SIDEBAR SETTINGS ---
st.sidebar.header("Security & Tools")
enable_protection = st.sidebar.checkbox("Enable Password Protection", value=True)
custom_password = "Smart_File_Lock"
if enable_protection:
    custom_password = st.sidebar.text_input("Set Sheet Password", value="Smart_File_Lock", type="password")

# --- TABS SETUP ---
tab1, tab2 = st.tabs(["🚀 Generator", "📖 User Manual"])

with tab1:
    st.markdown("Upload your files below to generate the distribution.")
    
    # --- FILE UPLOADERS ---
    col1, col2 = st.columns(2)
    with col1:
        items_file = st.file_uploader("Upload Participants List (Class List)", type=['xlsx'])
    with col2:
        template_file = st.file_uploader("Upload Team Names (With Class Label)", type=['xlsx'])

    if items_file and template_file:
        if st.button("🚀 Generate Distribution"):
            try:
                # 1. Load Data
                df_items = pd.read_excel(items_file)
                group_headers = sorted([c for c in df_items.columns if "Unnamed" not in str(c)])
                
                wb_tmpl = load_workbook(template_file, data_only=True)
                ws_tmpl = wb_tmpl.worksheets[0]
                
                # Create a map of Column A (Cleaned) to Row Number
                first_col_map = {re.sub(r'[^a-zA-Z0-9]', '', str(ws_tmpl.cell(row=r, column=1).value)).lower(): r 
                                 for r in range(1, ws_tmpl.max_row + 1) if ws_tmpl.cell(row=r, column=1).value}

                # 2. Setup Output Workbook
                wb_out = Workbook()
                summary_sheet = wb_out.active
                summary_sheet.title = "Distribution Summary"
                
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                summary_sheet.cell(row=1, column=1, value=f"Generated on: {now}").font = Font(italic=True)
                
                if enable_protection:
                    summary_sheet.cell(row=2, column=1, value="Note: Sheets are password protected.").font = Font(size=9, color="FF0000")
                
                # Styles
                summary_header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                summary_text_font = Font(color="FFFFFF", bold=True)
                team_header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                sn_column_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

                headers = ["Group Header", "Status / Row", "Teams Count", "Total Items"]
                for idx, text in enumerate(headers, 1):
                    cell = summary_sheet.cell(row=4, column=idx, value=text)
                    cell.font, cell.fill, cell.alignment = summary_text_font, summary_header_fill, Alignment(horizontal="center")

                summary_row, sheets_created = 5, 0
                browser_summary_data = []
                
                # Grand Totals Trackers
                total_teams_acc = 0
                total_items_acc = 0

                # 3. Process Logic
                for header in group_headers:
                    target = re.sub(r'[^a-zA-Z0-9]', '', str(header)).lower()
                    if target not in first_col_map:
                        summary_sheet.cell(row=summary_row, column=1, value=header)
                        summary_sheet.cell(row=summary_row, column=2, value="Skipped: No match")
                        summary_row += 1
                        browser_summary_data.append({"Group": header, "Teams": 0, "Items": 0, "Status": "❌ Skipped"})
                        continue
                    
                    anchor_r = first_col_map[target]
                    
                    # Clean items logic
                    items_raw = df_items[header].dropna().tolist()
                    items = [str(i)[:-2] if str(i).endswith('.0') else str(i).strip() for i in items_raw]
                    random.shuffle(items)

                    # Extract matrix teams
                    matrix_blocks, curr_r, team_count = [], anchor_r, 0
                    while curr_r <= ws_tmpl.max_row:
                        row_teams = []
                        for c in range(2, ws_tmpl.max_column + 1, 2):
                            s_val, n_val = ws_tmpl.cell(row=curr_r, column=c).value, ws_tmpl.cell(row=curr_r, column=c+1).value
                            if s_val is not None and n_val is not None:
                                row_teams.append({'s': s_val, 'n': str(n_val).strip(), 'col': c})
                                team_count += 1
                        if not row_teams:
                            if matrix_blocks: break
                        else: matrix_blocks.append(row_teams)
                        curr_r += 1

                    if not items or not matrix_blocks:
                        summary_sheet.cell(row=summary_row, column=1, value=header)
                        summary_sheet.cell(row=summary_row, column=2, value=f"Row {anchor_r} (Missing Matrix)")
                        summary_row += 1
                        browser_summary_data.append({"Group": header, "Teams": team_count, "Items": len(items), "Status": "⚠️ Empty"})
                        continue

                    # Create Class Sheet
                    ws_out = wb_out.create_sheet(title=str(header)[:30])
                    sheets_created += 1
                    if enable_protection: ws_out.protection.set_password(custom_password)

                    # Update Excel Summary Data
                    summary_sheet.cell(row=summary_row, column=1, value=header)
                    summary_sheet.cell(row=summary_row, column=2, value=f"Row {anchor_r}")
                    summary_sheet.cell(row=summary_row, column=3, value=team_count)
                    summary_sheet.cell(row=summary_row, column=4, value=len(items))
                    
                    total_teams_acc += team_count
                    total_items_acc += len(items)
                    summary_row += 1
                    browser_summary_data.append({"Group": header, "Teams": team_count, "Items": len(items), "Status": "✅ Success"})

                    # Distribute Items with AUTO-WIDTH LOGIC Restore
                    max_widths, sn_width = {}, 11.9
                    ws_out.column_dimensions['A'].width = sn_width
                    flat_teams = [t for block in matrix_blocks for t in block]
                    avg, rem = divmod(len(items), len(flat_teams))
                    item_ptr, out_r = 1, 1

                    for block in matrix_blocks:
                        max_block_h = 0
                        for team in block:
                            # S/N Column
                            col_let_sn = ws_out.cell(row=1, column=team['col']).column_letter
                            ws_out.column_dimensions[col_let_sn].width = sn_width
                            
                            # Data Column Header (Team Name) Width
                            col_let_data = ws_out.cell(row=1, column=team['col']+1).column_letter
                            max_widths[col_let_data] = max(max_widths.get(col_let_data, 0), len(str(team['n'])))

                            # Set Headers
                            c1, c2 = ws_out.cell(row=out_r, column=team['col'], value=team['s']), ws_out.cell(row=out_r, column=team['col']+1, value=team['n'])
                            for cell in [c1, c2]:
                                cell.font, cell.fill, cell.border = Font(bold=True), team_header_fill, no_border

                        for team in block:
                            t_idx = flat_teams.index(team)
                            count = avg + (1 if t_idx < rem else 0)
                            max_block_h = max(max_block_h, count)
                            for i in range(1, count + 1):
                                if item_ptr <= len(items):
                                    val = items[item_ptr-1]
                                    c_idx, c_val = ws_out.cell(row=out_r + i, column=team['col'], value=i), ws_out.cell(row=out_r + i, column=team['col']+1, value=val)
                                    c_idx.fill, c_idx.border, c_val.border = sn_column_fill, no_border, no_border
                                    c_val.number_format = '@'
                                    
                                    # Update Data Width for Long Participant Names
                                    col_let_data = ws_out.cell(row=1, column=team['col']+1).column_letter
                                    max_widths[col_let_data] = max(max_widths.get(col_let_data, 0), len(str(val)))
                                    
                                    item_ptr += 1
                        out_r += (max_block_h + 3)

                    # Apply Calculated Column Widths to the Sheet
                    for col_let, length in max_widths.items():
                        ws_out.column_dimensions[col_let].width = length + 3

                # --- 4. ADD GRAND TOTALS TO EXCEL SUMMARY ---
                sum_label = summary_sheet.cell(row=summary_row, column=1, value="GRAND TOTAL")
                sum_teams = summary_sheet.cell(row=summary_row, column=3, value=total_teams_acc)
                sum_items = summary_sheet.cell(row=summary_row, column=4, value=total_items_acc)
                for cell in [sum_label, sum_teams, sum_items]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # Excel Final Clean
                if enable_protection: summary_sheet.protection.set_password(custom_password)
                for col in ['A', 'B', 'C', 'D']: summary_sheet.column_dimensions[col].width = 25

                # Save for search feature
                st.session_state['active_wb'] = wb_out

                # --- 5. BROWSER DASHBOARD ---
                st.divider()
                sum_df = pd.DataFrame(browser_summary_data)
                
                m1, m2, m3 = st.columns(3)
                m1.metric("Total Classes", len(sum_df))
                m2.metric("Total Items (Grand Total)", total_items_acc)
                m3.metric("Total Teams (Grand Total)", total_teams_acc)

                left, right = st.columns([1, 1])
                with left:
                    st.subheader("📋 Browser Summary")
                    st.dataframe(
                        sum_df.style.set_properties(**{'background-color': '#F9F9F9', 'color': '#2C3E50'})
                        .set_table_styles([{'selector': 'th', 'props': [('background-color', '#2C3E50'), ('color', 'white')]}])
                        .background_gradient(cmap="Blues", subset=["Items"]),
                        use_container_width=True, hide_index=True
                    )
                with right:
                    st.subheader("🥧 Distribution Composition")
                    fig = px.pie(sum_df[sum_df["Items"] > 0], values='Items', names='Group', hole=0.4, color_discrete_sequence=px.colors.qualitative.Prism)
                    fig.update_layout(margin=dict(t=0, b=0, l=0, r=0))
                    st.plotly_chart(fig, use_container_width=True)

                output = io.BytesIO()
                wb_out.save(output)
                st.success("✅ Distribution Ready!")
                st.download_button(label="📥 Download Smart Batcher Excel", data=output.getvalue(), file_name=f"Smart_Batcher_{datetime.now().strftime('%H%M%S')}.xlsx")

            except Exception as e:
                st.error(f"Error: {e}")

# --- SEARCH FEATURE IN SIDEBAR ---
if 'active_wb' in st.session_state:
    st.sidebar.divider()
    st.sidebar.subheader("🔍 Find Participant")
    search_query = st.sidebar.text_input("Enter name to search:").strip().lower()
    
    if search_query:
        matches = []
        wb = st.session_state['active_wb']
        for sheetname in wb.sheetnames:
            if sheetname == "Distribution Summary": continue
            ws = wb[sheetname]
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if val and search_query in str(val).lower():
                        matches.append({"Name": val, "Group": sheetname})
        
        if matches:
            for m in matches:
                st.sidebar.success(f"**Found!**\n\n**Name:** {m['Name']}\n\n**Class:** {m['Group']}")
        else:
            st.sidebar.warning("No participant found.")

with tab2:
    st.header("📘 Smart Batcher User Manual")
    
    st.subheader("1. How it Works")
    st.write("Smart Batcher automates randomization of participants into team matrix structures while ensuring even distribution.")

    col_m1, col_m2 = st.columns(2)
    with col_m1:
        st.info("**File A: Participants List**")
        st.write("- **Format:** .xlsx\n- **Headers:** Top row = Class Names.\n- **Data:** Names listed under headers.")
    with col_m2:
        st.info("**File B: Team Template**")
        st.write("- **Format:** .xlsx\n- **Column A:** Class Names (must match File A).\n- **Columns B, C, etc:** Team names/numbers.")

    st.divider()
    st.subheader("2. Step-by-Step Guide")
    st.markdown("1. **Set Security:** Use the sidebar to set passwords.\n2. **Upload Files:** Drop your Excel files in the 'Generator' tab.\n3. **Generate:** Process the randomization.\n4. **Download:** Export your auto-formatted, protected Excel file.")

    try:
        with open("User_Manual.pdf", "rb") as f:
            st.download_button("📥 Download PDF Manual", f, "Smart_Batcher_Manual.pdf")
    except:
        st.caption("PDF Manual file not found in local directory.")
